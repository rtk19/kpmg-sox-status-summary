from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime
from pathlib import Path

app = Flask(__name__)

ROOT_DIR = Path(__file__).resolve().parent.parent
PUBLIC_DIR = ROOT_DIR / "public"
SRC_ASSETS_DIR = ROOT_DIR / "src" / "assets"

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

STATUS_HEADERS = {"מסקנה", "בקרה", "סטטוס", "status", "conclusion", "control", "results"}
KEY_HEADERS = {"מפתח", "בקרת מפתח", "key", "key control"}
PROCESS_HEADERS = {"תהליך", "process", "process name", "שם התהליך"}


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def is_status_header(text: str) -> bool:
    t = clean(text).lower()
    return t in STATUS_HEADERS or t.startswith("מסקנה") or t.startswith("conclusion")


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in vals if v]
        if len(non_empty) >= 2 and any(is_status_header(v) for v in non_empty):
            return r
    return None


def process_workbooks(file_streams):
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    summary = out_wb.create_sheet("סיכום")
    summary.append(["גיליון", "שם התהליך", "סבב", "סה\"כ בקרות", "בקרת מפתח", "סה\"כ סטטוסים", "אחוז ביצוע"])

    for fs, filename in file_streams:
        wb = load_workbook(fs)
        for ws in wb.worksheets:
            new_ws = out_wb.create_sheet(ws.title[:31])
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            hr = find_header_row(ws)
            if not hr:
                continue

            headers = [clean(ws.cell(hr, c).value).lower() for c in range(1, ws.max_column + 1)]
            status_cols = [i+1 for i,h in enumerate(headers) if is_status_header(h)]
            process_col = next((i+1 for i,h in enumerate(headers) if h in PROCESS_HEADERS), None)
            key_col = next((i+1 for i,h in enumerate(headers) if h in KEY_HEADERS), None)

            for idx, scol in enumerate(status_cols, start=1):
                process_map = {}
                for r in range(hr+1, ws.max_row+1):
                    if not any(clean(ws.cell(r, x).value) for x in (1,2,3) if x <= ws.max_column):
                        break
                    proc = clean(ws.cell(r, process_col).value) if process_col else "כללי"
                    if not proc:
                        proc = "כללי"
                    st = clean(ws.cell(r, scol).value)
                    key = clean(ws.cell(r, key_col).value).lower() if key_col else ""

                    rec = process_map.setdefault(proc, {"total":0, "key":0, "status":0})
                    rec["total"] += 1
                    if key in {"yes", "y", "כן"}:
                        rec["key"] += 1
                    if st:
                        rec["status"] += 1

                for proc, rec in process_map.items():
                    pct = round((rec["status"]/rec["total"])*100, 1) if rec["total"] else 0
                    summary.append([ws.title, proc, idx, rec["total"], rec["key"], rec["status"], pct])

    bio = BytesIO()
    out_wb.save(bio)
    bio.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return bio, f"SOX_Status_Summary_{ts}.xlsx"


@app.get('/')
def index():
    return render_template('index.html')



@app.get('/fonts/<path:filename>')
def fonts(filename):
    return send_from_directory(PUBLIC_DIR / 'fonts', filename)


@app.get('/assets/<path:filename>')
def assets(filename):
    return send_from_directory(SRC_ASSETS_DIR, filename)


@app.post('/process')
def process():
    files = request.files.getlist('files')
    if not files:
        return jsonify({"error":"לא נבחרו קבצים"}), 400
    bad = [f.filename for f in files if not f.filename.lower().endswith('.xlsx')]
    if bad:
        return jsonify({"error":"יש להעלות קבצים מסוג xlsx בלבד"}), 400

    try:
        out, fname = process_workbooks([(f.stream, f.filename) for f in files])
        return send_file(out, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error":str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
