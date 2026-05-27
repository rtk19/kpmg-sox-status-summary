from __future__ import annotations
from io import BytesIO
from typing import Iterable
from openpyxl import load_workbook, Workbook


def process_sox_files(file_streams: Iterable[tuple[str, bytes]]) -> tuple[bytes, str]:
    """Combine uploaded xlsx files into a single workbook and preserve sheets.

    This is a Python backend replacement for the previous browser-side processing.
    """
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for original_name, raw in file_streams:
        wb = load_workbook(filename=BytesIO(raw), data_only=False)
        base = original_name.rsplit('.', 1)[0][:18]
        for ws in wb.worksheets:
            title = f"{base}_{ws.title}"[:31]
            # ensure uniqueness
            if title in out_wb.sheetnames:
                i = 2
                cand = f"{title[:28]}_{i}"
                while cand in out_wb.sheetnames:
                    i += 1
                    cand = f"{title[:28]}_{i}"
                title = cand
            new_ws = out_wb.create_sheet(title=title)
            for row in ws.iter_rows():
                for cell in row:
                    new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

    if not out_wb.sheetnames:
        out_wb.create_sheet("summary")

    buffer = BytesIO()
    out_wb.save(buffer)
    return buffer.getvalue(), "sox_status_summary.xlsx"
