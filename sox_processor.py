from __future__ import annotations

import io
import json
import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


COLOR_HEADER_BG = "4F81BD"
COLOR_HEADER_FG = "FFFFFF"
COLOR_TOTAL_BG = "D9E1F2"
COLOR_ROW_ALT = "EEF3FA"
COLOR_ROW_NORMAL = "FFFFFF"

EXCLUDED_STATUS_VALUES = {"", "לא מפתח", "not key", "IT", "it", "annual", "שנתי"}

STATUS_HEADER_CANDIDATES = {
    "מסקנה",
    "בקרה",
    "סטטוס",
    "מסקנה סבב א'",
    "מסקנה סבב ב'",
    "מסקנה סבב א",
    "מסקנה סבב ב",
    "status",
    "conclusion",
    "control",
    "results",
}

ENGLISH_STATUS_HEADERS = {"status", "conclusion", "control", "results"}

HEBREW_LETTER_VALUES = {
    "א": 1,
    "ב": 2,
    "ג": 3,
    "ד": 4,
    "ה": 5,
    "ו": 6,
    "ז": 7,
    "ח": 8,
    "ט": 9,
    "י": 10,
    "כ": 20,
    "ל": 30,
    "מ": 40,
    "נ": 50,
    "ס": 60,
    "ע": 70,
    "פ": 80,
    "צ": 90,
    "ק": 100,
    "ר": 200,
    "ש": 300,
    "ת": 400,
}

KEY_PRIORITY_1 = {"מפתח", "key control"}
KEY_PRIORITY_2 = {"בקרת מפתח", "key"}

ID_HEADER_PRIORITY_EXACT = {
    "מס",
    "מס'",
    "מספר",
    "מס בקרה",
    "מס' בקרה",
    "מספר בקרה",
    "control number",
    "control no",
    "id",
    "no",
    "no.",
    ".no",
}

ID_HEADER_CONTAINS = [
    "מס",
    "מס'",
    "מספר",
    "control number",
    "control no",
    "id",
    "no",
    "no.",
    ".no",
]

NON_ID_HEADERS_EXACT = {"מפתח", "בקרת מפתח", "מסקנה", "סטטוס", "בקרה", "key control"}

PROCESS_HEADER_PRIMARY = {"process", "תהליך"}
PROCESS_HEADER_SECONDARY = {"process name", "שם התהליך"}
PROCESS_HEADER_EXCLUDE_TOKENS = ["sub", "sub-", "sub_", "משנה", "תת"]

Lang = Literal["he", "en"]

LABELS: dict[Lang, dict[str, str]] = {
    "he": {
        "sheet_name": "גיליון",
        "process_name": "שם התהליך",
        "round": "סבב",
        "total_controls": 'סה"כ בקרות',
        "key_control": "בקרת מפתח",
        "total_statuses": 'סה"כ סטטוסים',
        "it_controls": "בקרות IT",
        "annual": "שנתי",
        "exec_pct": "אחוז ביצוע",
        "total_row": 'סה"כ',
    },
    "en": {
        "sheet_name": "Sheet",
        "process_name": "Process Name",
        "round": "Round",
        "total_controls": "Total Controls",
        "key_control": "Key Control",
        "total_statuses": "Total Statuses",
        "it_controls": "IT Controls",
        "annual": "Annual",
        "exec_pct": "Execution %",
        "total_row": "Total",
    },
}


class SoxBusinessStop(Exception):
    pass


class SoxClarificationRequired(SoxBusinessStop):
    pass


class SoxStructureNotIdentified(SoxBusinessStop):
    pass


@dataclass
class InputFile:
    name: str
    data: bytes


@dataclass
class ProcessResult:
    data: bytes
    filename: str


@dataclass
class SheetConfig:
    start_row: int
    end_row: int
    id_col: str | None
    key_col: str | None
    process_col: str | None
    status_cols: list[str]
    round_numbers: list[int]


@dataclass
class SummaryRow:
    sheet: str
    process: str | None
    status_index: int
    cfg: SheetConfig


def is_status_header(text: str) -> bool:
    if not text:
        return False
    if text in STATUS_HEADER_CANDIDATES:
        return True
    if re.match(r"^מסקנה(\s|$)", text):
        return True
    if re.match(r"^conclusion(\s|$)", text):
        return True
    return False


def is_english_status_header(text: str) -> bool:
    if text in ENGLISH_STATUS_HEADERS:
        return True
    return bool(re.match(r"^conclusion(\s|$)", text))


def is_hebrew_status_header(text: str) -> bool:
    if not text:
        return False
    if is_english_status_header(text):
        return False
    return is_status_header(text)


def hebrew_letters_to_number(letters: str) -> int | None:
    total = 0
    for ch in letters:
        value = HEBREW_LETTER_VALUES.get(ch)
        if not value:
            return None
        total += value
    return total if total > 0 else None


def parse_round_number(text: str) -> int | None:
    if not text:
        return None
    digit_match = re.search(r"(\d+)", text)
    if digit_match:
        return int(digit_match.group(1))
    heb_match = re.search(r"סבב\s+([\u05d0-\u05ea]+)'?", text)
    if heb_match:
        return hebrew_letters_to_number(heb_match.group(1).replace("'", ""))
    return None


def strip_invisible(value: str) -> str:
    return re.sub(r"[\u00a0\u200b\u200c\u200d\ufeff]", " ", value).strip()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return strip_invisible(str(value))


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).lower()


def is_empty(value: Any) -> bool:
    return clean_text(value) == ""


def escape_excel(value: str) -> str:
    return value.replace('"', '""')


def looks_like_header_cell(text: str) -> bool:
    if text == "":
        return False
    return any(ch.isalpha() for ch in text)


def get_max_row(ws: Worksheet) -> int:
    return ws.max_row or 0


def get_max_col(ws: Worksheet) -> int:
    max_col = ws.max_column or 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column > max_col:
                max_col = cell.column
    return max_col


def cell_value(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def row_values(ws: Worksheet, row: int, max_col: int) -> list[str]:
    return [clean_text(cell_value(ws, row, col)) for col in range(1, max_col + 1)]


def count_non_empty_in_row(ws: Worksheet, row: int, max_col: int) -> int:
    return len([value for value in row_values(ws, row, max_col) if value != ""])


def normalize_status_values(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text in EXCLUDED_STATUS_VALUES:
            continue
        if text not in seen:
            seen.add(text)
            out.append(text)
    return out


def score_header_row(ws: Worksheet, row: int, max_col: int) -> int:
    values = row_values(ws, row, max_col)
    non_empty = [value for value in values if value != ""]
    if len(non_empty) < 2:
        return -1

    header_like = len([value for value in non_empty if looks_like_header_cell(value)])
    exact_status = len([value for value in non_empty if is_status_header(clean_header(value))])
    key_hits = len(
        [
            value
            for value in non_empty
            if clean_header(value) in KEY_PRIORITY_1
            or clean_header(value) in KEY_PRIORITY_2
            or ("מפתח" in clean_header(value) and "לא מפתח" not in clean_header(value))
        ]
    )
    id_hits = 0
    for value in non_empty:
        text = clean_header(value)
        if text in ID_HEADER_PRIORITY_EXACT or any(tok.lower() in text.lower() for tok in ID_HEADER_CONTAINS):
            id_hits += 1

    max_row = get_max_row(ws)
    next_row_non_empty = count_non_empty_in_row(ws, row + 1, max_col) if row < max_row else 0

    return len(non_empty) * 3 + header_like * 2 + exact_status * 6 + key_hits * 5 + id_hits * 4 + min(
        next_row_non_empty, 10
    )


def find_header_row(ws: Worksheet) -> int:
    max_col = get_max_col(ws)
    max_row = get_max_row(ws)
    scan_until = min(max_row, 30)
    best_row: int | None = None
    best_score = -1
    for row in range(1, scan_until + 1):
        score = score_header_row(ws, row, max_col)
        if score > best_score:
            best_score = score
            best_row = row
    if best_row is None or best_score < 0:
        raise SoxStructureNotIdentified(
            f"לא זוהתה שורת כותרת בגיליון '{ws.title}'.\n"
            "איך לתקן: ודא שב-30 השורות הראשונות קיימת שורה אחת לפחות עם כותרות עמודות (לפחות שתי עמודות עם טקסט).\n"
            "דוגמאות לכותרות מזוהות: מסקנה, סטטוס, בקרה, מסקנה סבב א', מפתח, בקרת מפתח, מס' בקרה, תהליך (או באנגלית: status, conclusion, control, key control, control number, process)."
        )
    return best_row


def find_table_bounds(ws: Worksheet, header_row: int) -> int:
    max_row = get_max_row(ws)
    first_data_row = header_row + 1
    no_data_msg = (
        f"לא נמצאו שורות נתונים מתחת לכותרת בגיליון '{ws.title}'.\n"
        "איך לתקן: ודא שמתחת לשורת הכותרת קיימת לפחות שורה אחת עם ערך באחת מ-3 העמודות הראשונות (A/B/C). "
        'בלוקים של הערות (כמו "Control Objectives") שמופיעים רק מעמודה D ואילך לא נחשבים כשורות נתונים.'
    )
    if first_data_row > max_row:
        raise SoxStructureNotIdentified(no_data_msg)

    for row in range(first_data_row, max_row + 1):
        a_empty = is_empty(cell_value(ws, row, 1))
        b_empty = is_empty(cell_value(ws, row, 2))
        c_empty = is_empty(cell_value(ws, row, 3))
        if a_empty and b_empty and c_empty:
            if row == first_data_row:
                raise SoxStructureNotIdentified(no_data_msg)
            return row - 1
    return max_row


def find_status_columns(ws: Worksheet, header_row: int) -> tuple[list[str], list[int | None]]:
    letters: list[str] = []
    round_numbers: list[int | None] = []
    max_col = get_max_col(ws)
    for col in range(1, max_col + 1):
        text = clean_header(cell_value(ws, header_row, col))
        if is_status_header(text):
            letters.append(get_column_letter(col))
            round_numbers.append(parse_round_number(text))
    return letters, round_numbers


def find_key_column(ws: Worksheet, header_row: int, overrides: dict[str, str]) -> str | None:
    if ws.title in overrides:
        return overrides[ws.title]

    headers: list[tuple[str, str]] = []
    max_col = get_max_col(ws)
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        text = clean_header(cell_value(ws, header_row, col))
        if text != "":
            headers.append((letter, text))

    p1 = [(letter, text) for letter, text in headers if text in KEY_PRIORITY_1]
    p2 = [(letter, text) for letter, text in headers if text in KEY_PRIORITY_2]
    p3 = [
        (letter, text)
        for letter, text in headers
        if "מפתח" in text and "לא מפתח" not in text and text not in KEY_PRIORITY_1 and text not in KEY_PRIORITY_2
    ]

    for matches in [p1, p2, p3]:
        if len(matches) == 1:
            return matches[0][0]
        if len(matches) > 1:
            lines = [f"נמצאו מספר עמודות אפשריות לעמודת המפתח בגיליון '{ws.title}':"]
            for letter, text in matches:
                lines.append(f'  • עמודה {letter} — "{text}"')
            lines.extend(["", "איך לתקן: השאר רק כותרת אחת שמכילה 'מפתח' / 'key control', או שנה את שמות שאר העמודות."])
            raise SoxClarificationRequired("\n".join(lines))
    return None


def find_id_column(ws: Worksheet, header_row: int, status_cols: set[str], key_col: str | None) -> str | None:
    headers: list[dict[str, Any]] = []
    max_col = get_max_col(ws)
    for col in range(1, max_col + 1):
        headers.append({"c": col, "letter": get_column_letter(col), "text": clean_header(cell_value(ws, header_row, col))})

    exact_lower = {item.lower() for item in ID_HEADER_PRIORITY_EXACT}
    exact = [header for header in headers if header["text"].lower() in exact_lower]
    if exact:
        return exact[0]["letter"]

    contains: list[dict[str, Any]] = []
    for header in headers:
        text = header["text"]
        if text == "":
            continue
        if header["letter"] in status_cols:
            continue
        if key_col and header["letter"] == key_col:
            continue
        if text in NON_ID_HEADERS_EXACT:
            continue
        lower = text.lower()
        if any(tok.lower() in lower for tok in ID_HEADER_CONTAINS):
            contains.append(header)
    if contains:
        return contains[0]["letter"]
    return None


def find_process_column(ws: Worksheet, header_row: int) -> str | None:
    max_col = get_max_col(ws)
    primary: list[str] = []
    secondary: list[str] = []
    for col in range(1, max_col + 1):
        text = clean_header(cell_value(ws, header_row, col))
        if text == "":
            continue
        if any(tok in text for tok in PROCESS_HEADER_EXCLUDE_TOKENS):
            continue
        if text in PROCESS_HEADER_PRIMARY:
            primary.append(get_column_letter(col))
        elif text in PROCESS_HEADER_SECONDARY:
            secondary.append(get_column_letter(col))
    if primary:
        return primary[0]
    if secondary:
        return secondary[0]
    return None


def visible_worksheets(wb: Workbook) -> list[Worksheet]:
    return [ws for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]


def detect_language(wb: Workbook, cfg: dict[str, SheetConfig]) -> Lang:
    for name, config in cfg.items():
        if not config.status_cols:
            continue
        ws = wb[name] if name in wb.sheetnames else None
        if not ws:
            continue
        max_col = get_max_col(ws)
        for col in range(1, max_col + 1):
            text = clean_header(cell_value(ws, config.start_row, col))
            if is_english_status_header(text):
                return "en"
            if is_hebrew_status_header(text):
                return "he"
    return "he"


def analyze_workbook(wb: Workbook, manual_overrides: dict[str, str]) -> tuple[dict[str, SheetConfig], Lang]:
    visible = visible_worksheets(wb)
    empty_file_msg = (
        "לא נמצא בקובץ אף גיליון גלוי עם נתונים.\n"
        "איך לתקן: ודא שהקובץ מכיל לפחות גיליון אחד שאינו מוסתר (Hidden) ושיש בו טבלת בקרות עם שורת כותרת וערכים."
    )
    if not visible:
        raise SoxStructureNotIdentified(empty_file_msg)

    sheet_config: dict[str, SheetConfig] = {}
    skipped: list[dict[str, str]] = []

    for ws in visible:
        max_row = get_max_row(ws)
        max_col = get_max_col(ws)
        if max_row <= 1 and max_col <= 1 and is_empty(cell_value(ws, 1, 1)):
            continue

        try:
            header_row = find_header_row(ws)
            status_cols, parsed_round_numbers = find_status_columns(ws, header_row)
            if len(status_cols) == 0:
                skipped.append({"name": ws.title, "reason": "לא נמצאה עמודת מסקנה/סטטוס בגיליון זה."})
                continue
            end_row = find_table_bounds(ws, header_row)
            key_col = find_key_column(ws, header_row, manual_overrides)
            id_col = find_id_column(ws, header_row, set(status_cols), key_col)
            process_col = find_process_column(ws, header_row)

            sheet_config[ws.title] = SheetConfig(
                start_row=header_row,
                end_row=end_row,
                id_col=id_col,
                key_col=key_col,
                process_col=process_col,
                status_cols=status_cols,
                round_numbers=[n or 0 for n in parsed_round_numbers],
            )
        except SoxClarificationRequired:
            raise
        except SoxBusinessStop as exc:
            first_line = str(exc).split("\n")[0]
            skipped.append({"name": ws.title, "reason": first_line})
            continue

    if not sheet_config:
        skipped_list = ""
        if skipped:
            skipped_list = "\n\nגיליונות שנדלגו:\n" + "\n".join(
                [f"  • '{item['name']}' — {item['reason']}" for item in skipped]
            )
        raise SoxStructureNotIdentified(
            "לא נמצא בקובץ אף גיליון תקין לעיבוד.\n"
            "איך לתקן: ודא שלפחות גיליון אחד מכיל שורת כותרת עם עמודת מסקנה/סטטוס "
            "(מסקנה, סטטוס, בקרה, status, conclusion, control, results) ושורות נתונים מתחתיה."
            + skipped_list
        )

    if skipped:
        print(
            "[SOX] גיליונות שנדלגו בעיבוד:\n"
            + "\n".join([f"  • '{item['name']}' — {item['reason']}" for item in skipped])
        )

    lang = detect_language(wb, sheet_config)
    for config in sheet_config.values():
        config.round_numbers = [idx + 1 for idx, _ in enumerate(config.status_cols)]
    return sheet_config, lang


def extract_all_dynamic_status_values(wb: Workbook, sheet_config: dict[str, SheetConfig]) -> list[str]:
    values: list[str] = []
    for name, cfg in sheet_config.items():
        ws = wb[name] if name in wb.sheetnames else None
        if not ws:
            continue
        for col in cfg.status_cols:
            col_idx = column_index_from_string(col)
            for row in range(cfg.start_row + 1, cfg.end_row + 1):
                text = clean_text(cell_value(ws, row, col_idx))
                if text == "":
                    continue
                if text in EXCLUDED_STATUS_VALUES:
                    continue
                values.append(text)
    return normalize_status_values(values)


def get_process_values(ws: Worksheet, cfg: SheetConfig) -> list[str]:
    if not cfg.process_col:
        return []
    col_idx = column_index_from_string(cfg.process_col)
    seen: set[str] = set()
    out: list[str] = []
    for row in range(cfg.start_row + 1, cfg.end_row + 1):
        text = clean_text(cell_value(ws, row, col_idx))
        if text == "":
            continue
        if text not in seen:
            seen.add(text)
            out.append(text)
    return out


def get_max_status_count(sheet_config: dict[str, SheetConfig]) -> int:
    return max([len(cfg.status_cols) for cfg in sheet_config.values()] or [0])


THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE)


def style_header_cell(cell: Cell) -> None:
    cell.font = Font(bold=True, color=COLOR_HEADER_FG)
    cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def set_formula(cell: Cell, formula: str) -> None:
    compact = re.sub(r"\s+", " ", formula).strip()
    if not compact.startswith("="):
        compact = "=" + compact
    cell.value = compact


SUMMARY_SHEET_NAME: dict[Lang, str] = {"he": "סיכום", "en": "Summary"}


def build_summary_rows(wb: Workbook, sheet_config: dict[str, SheetConfig]) -> list[SummaryRow]:
    rows: list[SummaryRow] = []
    for name, cfg in sheet_config.items():
        ws = wb[name] if name in wb.sheetnames else None
        if not ws:
            continue
        processes = get_process_values(ws, cfg)
        process_col_idx = column_index_from_string(cfg.process_col) if cfg.process_col else 0

        def has_any_status_value(status_col: str, process: str | None) -> bool:
            col_idx = column_index_from_string(status_col)
            for row in range(cfg.start_row + 1, cfg.end_row + 1):
                if process and process_col_idx:
                    proc_text = clean_text(cell_value(ws, row, process_col_idx))
                    if proc_text != process:
                        continue
                text = clean_text(cell_value(ws, row, col_idx))
                if text != "":
                    return True
            return False

        for idx, status_col in enumerate(cfg.status_cols):
            if processes:
                for process in processes:
                    if not has_any_status_value(status_col, process):
                        continue
                    rows.append(SummaryRow(sheet=name, process=process, status_index=idx, cfg=cfg))
            else:
                if not has_any_status_value(status_col, None):
                    continue
                rows.append(SummaryRow(sheet=name, process=None, status_index=idx, cfg=cfg))
    return rows


def create_summary_sheet(wb: Workbook, sheet_config: dict[str, SheetConfig], dynamic_values: list[str], lang: Lang) -> None:
    labels = LABELS[lang]
    filtered = normalize_status_values(dynamic_values)
    summary_rows = build_summary_rows(wb, sheet_config)

    max_rounds = get_max_status_count(sheet_config)
    show_round = max_rounds > 1
    show_process = any(cfg.process_col for cfg in sheet_config.values())

    sheet_name = SUMMARY_SHEET_NAME[lang]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = lang == "he"

    headers = [labels["sheet_name"]]
    if show_process:
        headers.append(labels["process_name"])
    if show_round:
        headers.append(labels["round"])
    headers.extend(
        [
            labels["total_controls"],
            labels["key_control"],
            *filtered,
            labels["total_statuses"],
            labels["it_controls"],
            labels["annual"],
            labels["exec_pct"],
        ]
    )

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        style_header_cell(cell)
    ws.row_dimensions[1].height = 30

    col = 1
    col_sheet = col
    col += 1
    col_process = col if show_process else 0
    if show_process:
        col += 1
    col_round = col if show_round else 0
    if show_round:
        col += 1
    col_total = col
    col += 1
    col_key = col
    col += 1
    col_status_start = col
    col += len(filtered)
    col_status_end = col_status_start + len(filtered) - 1 if filtered else col_status_start - 1
    col_row_sum = col
    col += 1
    col_it = col
    col += 1
    col_annual = col
    col += 1
    col_exec = col

    pass_idx = next((idx for idx, value in enumerate(filtered) if value.strip().lower() in {"עבר", "pass"}), -1)
    pass_col_letter = get_column_letter(col_status_start + pass_idx) if pass_idx >= 0 else None
    col_key_letter = get_column_letter(col_key)
    col_it_letter = get_column_letter(col_it)
    col_annual_letter = get_column_letter(col_annual)

    def proc_filter(cfg: SheetConfig, sheet: str, process: str | None) -> str:
        if not process or not cfg.process_col:
            return ""
        range_ref = f"'{sheet}'!{cfg.process_col}{cfg.start_row + 1}:{cfg.process_col}{cfg.end_row}"
        return f'(TRIM({range_ref})="{escape_excel(process)}")*'

    curr = 2
    wrote_data = False

    for summary_row in summary_rows:
        sheet = summary_row.sheet
        process = summary_row.process
        status_index = summary_row.status_index
        cfg = summary_row.cfg
        status_col = cfg.status_cols[status_index]
        start = cfg.start_row + 1
        end = cfg.end_row
        status_range = f"'{sheet}'!{status_col}{start}:{status_col}{end}"
        pf = proc_filter(cfg, sheet, process)

        ws.cell(row=curr, column=col_sheet).value = sheet
        if show_process:
            ws.cell(row=curr, column=col_process).value = process or ""
        if show_round:
            ws.cell(row=curr, column=col_round).value = cfg.round_numbers[status_index] or status_index + 1

        if cfg.id_col:
            id_range = f"'{sheet}'!{cfg.id_col}{start}:{cfg.id_col}{end}"
            set_formula(ws.cell(row=curr, column=col_total), f"=SUMPRODUCT({pf}(--ISNUMBER(--RIGHT({id_range},1))))")
        else:
            ws.cell(row=curr, column=col_total).value = 0

        if cfg.key_col:
            key_range = f"'{sheet}'!{cfg.key_col}{start}:{cfg.key_col}{end}"
            if pf:
                set_formula(
                    ws.cell(row=curr, column=col_key),
                    f'=SUMPRODUCT({pf}(({key_range}="כן")+({key_range}="Yes")+({key_range}="Y / כן")))',
                )
            else:
                set_formula(
                    ws.cell(row=curr, column=col_key),
                    f'=COUNTIF({key_range},"כן")+COUNTIF({key_range},"Yes")+COUNTIF({key_range},"Y / כן")',
                )
        else:
            ws.cell(row=curr, column=col_key).value = 0

        for idx, value in enumerate(filtered):
            set_formula(
                ws.cell(row=curr, column=col_status_start + idx),
                f'=SUMPRODUCT({pf}(--(TRIM({status_range})="{escape_excel(value)}")))',
            )

        if filtered:
            d_start = get_column_letter(col_status_start)
            d_end = get_column_letter(col_status_end)
            set_formula(ws.cell(row=curr, column=col_row_sum), f"=SUM({d_start}{curr}:{d_end}{curr})")
        else:
            ws.cell(row=curr, column=col_row_sum).value = 0

        set_formula(ws.cell(row=curr, column=col_it), f'=SUMPRODUCT({pf}(--ISNUMBER(SEARCH("IT",TRIM({status_range})))))')
        set_formula(
            ws.cell(row=curr, column=col_annual),
            f'=SUMPRODUCT({pf}(--((TRIM({status_range})="שנתי")+(TRIM({status_range})="Annual"))))',
        )

        if pass_col_letter:
            set_formula(
                ws.cell(row=curr, column=col_exec),
                f"=IFERROR({pass_col_letter}{curr}/MAX({pass_col_letter}{curr},{col_key_letter}{curr}-{col_it_letter}{curr}-{col_annual_letter}{curr}),0)",
            )
        else:
            ws.cell(row=curr, column=col_exec).value = 0

        fill_color = COLOR_ROW_ALT if curr % 2 == 0 else COLOR_ROW_NORMAL
        for c in range(1, col_exec + 1):
            cell = ws.cell(row=curr, column=c)
            cell.border = THIN_BORDER
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            align_boundary = col_process if show_process else col_sheet
            cell.alignment = Alignment(horizontal="right" if c <= align_boundary else "center", vertical="center")

        wrote_data = True
        curr += 1

    label_cell = ws.cell(row=curr, column=1)
    label_cell.value = labels["total_row"]
    label_cell.font = Font(bold=True)
    label_cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_BG)
    label_cell.border = THIN_BORDER
    label_cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(2, col_total):
        cell = ws.cell(row=curr, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_BG)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in range(col_total, col_exec):
        letter = get_column_letter(c)
        cell = ws.cell(row=curr, column=c)
        set_formula(cell, f"=SUM({letter}2:{letter}{curr - 1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_BG)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    final_exec = ws.cell(row=curr, column=col_exec)
    if pass_col_letter and wrote_data:
        set_formula(
            final_exec,
            f"=IFERROR({pass_col_letter}{curr}/MAX({pass_col_letter}{curr},{col_key_letter}{curr}-{col_it_letter}{curr}-{col_annual_letter}{curr}),0)",
        )
    else:
        final_exec.value = 0
    final_exec.font = Font(bold=True)
    final_exec.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_BG)
    final_exec.border = THIN_BORDER
    final_exec.alignment = Alignment(horizontal="center", vertical="center")
    final_exec.number_format = "0%"

    for row in range(2, curr + 1):
        ws.cell(row=row, column=col_exec).number_format = "0%"

    for c in range(1, col_exec + 1):
        max_len = 0
        for r in range(1, curr + 1):
            value = ws.cell(row=r, column=c).value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 30)


def sanitize_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", name)[:31]


def unique_sheet_name(base: str, used: set[str]) -> str:
    candidate = sanitize_sheet_name(base)
    idx = 2
    while candidate in used:
        suffix = f" ({idx})"
        idx += 1
        candidate = sanitize_sheet_name(base)[: 31 - len(suffix)] + suffix
    return candidate


def copy_sheet(src: Worksheet, dest_wb: Workbook, name: str) -> Worksheet:
    dest = dest_wb.create_sheet(name)
    dest.sheet_view.rightToLeft = src.sheet_view.rightToLeft
    dest.freeze_panes = src.freeze_panes

    for key, dimension in src.column_dimensions.items():
        dest.column_dimensions[key].width = dimension.width
        dest.column_dimensions[key].hidden = dimension.hidden

    for idx, dimension in src.row_dimensions.items():
        dest.row_dimensions[idx].height = dimension.height
        dest.row_dimensions[idx].hidden = dimension.hidden

    for row in src.iter_rows():
        for source_cell in row:
            target = dest.cell(row=source_cell.row, column=source_cell.column)
            target.value = source_cell.value
            if source_cell.has_style:
                target.font = copy(source_cell.font)
                target.fill = copy(source_cell.fill)
                target.border = copy(source_cell.border)
                target.alignment = copy(source_cell.alignment)
                target.number_format = source_cell.number_format
                target.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target.comment = copy(source_cell.comment)

    for merged_range in src.merged_cells.ranges:
        try:
            dest.merge_cells(str(merged_range))
        except Exception:
            pass

    return dest


def load_workbook_from_input(file: InputFile) -> Workbook:
    try:
        return load_workbook(io.BytesIO(file.data))
    except Exception as exc:
        msg = str(exc)
        if re.search(r"old \.xls|BIFF", msg, re.I):
            raise SoxBusinessStop(
                f"הקובץ '{file.name}' הוא בפורמט xls הישן ואינו נתמך.\n"
                'איך לתקן: פתחו אותו ב-Excel ובחרו "Save As" → סוג קובץ "Excel Workbook (.xlsx)", ואז העלו שוב.'
            ) from exc
        raise SoxBusinessStop(
            f"הקובץ '{file.name}' אינו קובץ xlsx תקין.\n"
            f"איך לתקן: ודא שסיומת הקובץ היא .xlsx ושהוא אינו פגום או מוגן בסיסמה. פרטים: {msg}"
        ) from exc


def new_empty_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def process_sox_file(file: InputFile, manual_overrides: dict[str, str] | None = None) -> ProcessResult:
    return process_sox_files([file], manual_overrides or {})


def process_sox_files(files: list[InputFile], manual_overrides: dict[str, str] | None = None) -> ProcessResult:
    if not files:
        raise SoxBusinessStop("לא נבחרו קבצים")

    manual_overrides = manual_overrides or {}
    combined = new_empty_workbook()
    used: set[str] = set()
    langs: set[Lang] = set()

    for file in files:
        wb = load_workbook_from_input(file)

        try:
            _, file_lang = analyze_workbook(wb, manual_overrides)
            langs.add(file_lang)
        except Exception:
            pass

        stem = re.sub(r"\.xlsx$", "", file.name, flags=re.I)
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            if re.match(r"^(status_\d+|סיכום|summary)$", ws.title, re.I):
                continue

            base = ws.title
            if len(files) > 1 and sanitize_sheet_name(base) in used:
                base = f"{stem} - {ws.title}"
            name = unique_sheet_name(base, used)
            used.add(name)
            copy_sheet(ws, combined, name)

    if not combined.worksheets:
        raise SoxStructureNotIdentified(
            "לא נמצא בקובץ אף גיליון גלוי עם נתונים.\n"
            "איך לתקן: ודא שהקובץ מכיל לפחות גיליון אחד שאינו מוסתר ושיש בו טבלת בקרות."
        )

    sheet_config, detected_lang = analyze_workbook(combined, manual_overrides)
    lang: Lang = "en" if len(langs) > 1 else detected_lang

    for ws in list(combined.worksheets):
        if re.match(r"^status_\d+$", ws.title, re.I):
            del combined[ws.title]

    dynamic_values = extract_all_dynamic_status_values(combined, sheet_config)
    create_summary_sheet(combined, sheet_config, dynamic_values, lang)

    out = io.BytesIO()
    combined.save(out)
    if len(files) == 1:
        filename = re.sub(r"\s+", "_", re.sub(r"\.xlsx$", "", files[0].name, flags=re.I)) + "_with_status.xlsx"
    else:
        filename = "sox_summary_of_controls_combined_with_status.xlsx"
    return ProcessResult(data=out.getvalue(), filename=filename)


def classify_error(err: Exception) -> str:
    if isinstance(err, SoxBusinessStop):
        return str(err)

    msg = str(err)
    if "No sheet" in msg or "Worksheet" in msg:
        return "הקובץ אינו מכיל גיליונות תקינים.\nאיך לתקן: ודא שהקובץ מכיל לפחות גיליון אחד גלוי עם נתונים."
    if "not an Excel file" in msg or "zip" in msg.lower():
        return "הקובץ אינו קובץ Excel תקין.\nאיך לתקן: ודא שהפורמט הוא .xlsx ושהקובץ אינו פגום."
    return f"שגיאה לא צפויה בעיבוד הקובץ.\nפרטים טכניים: {msg}"


def overrides_from_json(raw: str | None) -> dict[str, str]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): str(value) for key, value in parsed.items()}
